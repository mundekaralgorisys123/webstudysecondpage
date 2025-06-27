
from flask import Flask, render_template, request,send_file, jsonify
import os
import re
from collections import Counter
import openpyxl
from PIL import Image
from openpyxl.drawing.image import Image as XLImage
import pymssql
import tempfile
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
import openpyxl
from PIL import Image
from openpyxl.drawing.image import Image as XLImage
import pymssql
import tempfile
from urllib.parse import urlparse
from proxy import check_proxies
from scrapers.ernest_jones import handle_ernest_jones
from scrapers.shaneco import handle_shane_co
from scrapers.fhinds import handle_fhinds
from scrapers.gabriel import handle_gabriel
from scrapers.hsamuel import handle_h_samuel
from scrapers.kay import handle_kay
from scrapers.jared import handle_jared
from scrapers.tiffany import handle_tiffany
#==========================================#
from scrapers.kayoutlet import handle_kayoutlet
from scrapers.zales import handle_zales
from scrapers.helzberg import handle_helzberg
from scrapers.rosssimons import handle_rosssimons
from scrapers.peoplesjewellers import handle_peoplesjewellers
from scrapers.fraserhart import handle_fraserhart
from scrapers.fields import handle_fields
from scrapers.warrenjames import handle_warrenjames
from scrapers.goldsmiths import handle_goldsmiths
from scrapers.thediamondstore import handle_thediamondstore
from scrapers.prouds import handle_prouds
from scrapers.goldmark import handle_goldmark
from scrapers.anguscoote import handle_anguscoote
from scrapers.bash import handle_bash
from scrapers.shiels import handle_shiels
from scrapers.mazzucchellis import handle_mazzucchellis
from scrapers.hoskings import handle_hoskings
from scrapers.hardybrothers import handle_hardybrothers
from scrapers.zamels import handle_zamels
from scrapers.wallacebishop import handle_wallacebishop
from scrapers.bevilles import handle_bevilles
from scrapers.michaelhill import handle_michaelhill
from scrapers.apart import handle_apart
from scrapers.macys import handle_macys
from scrapers.jcpenney import handle_jcpenney
from scrapers.fredmeyer import handle_fredmeyer
from scrapers.beaverbrooks import handle_beaverbrooks
#############################################################################################################
            #stage 2
#############################################################################################################
from scrapers.finks import handle_finks
from scrapers.smilingrocks import handle_smilingrocks
from scrapers.bluenile import handle_bluenile
from scrapers.benbridge import handle_benbridge
from scrapers.hannoush import handle_hannoush
from scrapers.jcojewellery import handle_jcojewellery
from scrapers.diamonds import handle_77diamonds
from scrapers.reeds import handle_reeds
from scrapers.walmart import handle_walmart
#############################################################################################################
from scrapers.armansfinejewellery import handle_armansfinejewellery
from scrapers.jacquefinejewellery import handle_jacquefinejewellery
from scrapers.medleyjewellery import handle_medleyjewellery
from scrapers.cullenjewellery import handle_cullenjewellery
from scrapers.grahams import handle_grahams
from scrapers.larsenjewellery import handle_larsenjewellery
from scrapers.ddsdiamonds import handle_ddsdiamonds
from scrapers.garenjewellery import handle_garenjewellery
from scrapers.stefandiamonds import handle_stefandiamonds
from scrapers.goodstoneinc import handle_goodstoneinc
from scrapers.natashaschweitzer import handle_natasha
from scrapers.sarahandsebastian import handle_sarahandsebastian
from scrapers.moissanite import handle_moissanite
from scrapers.daimondcollection import handle_diamondcollection
from scrapers.cushlawhiting import handle_cushlawhiting
from scrapers.cerrone import handle_cerrone
from scrapers.briju import handle_briju
from scrapers.histoiredor import handle_histoiredor
from scrapers.marcorian import handle_marcorian
from scrapers.klenotyaurum import handle_klenotyaurum
from scrapers.stroilioro import handle_stroilioro
from scrapers.mariemass import handle_mariemass
from scrapers.mattioli import handle_mattioli
from scrapers.pomellato import handle_pomellato
from scrapers.dior import handle_dior
from scrapers.bonnie import handle_bonnie

########################################### 24/07 ################################################################## 
from scrapers.diamondsfactory import handle_diamondsfactory
from scrapers.davidmarshalllondon import handle_davidmarshalllondon
from scrapers.monicavinader import handle_monicavinader
from scrapers.boodles import handle_boodles
from scrapers.mariablack import handle_mariablack
from scrapers.londonjewelers import handle_londonjewelers
from scrapers.fernandojorge import handle_fernandojorge
from scrapers.pandora import handle_pandora
from scrapers.daisyjewellery import handle_daisyjewellery
from scrapers.missoma import handle_missoma
from scrapers.astleyclarke import handle_astleyclarke
from scrapers.edgeofember import handle_edgeofember
from scrapers.mateo import handle_mateo
from scrapers.bybonniejewelry import handle_bybonniejewelry
################################################ 25/04 ############################################################# 
from scrapers.tacori import handle_tacori
from scrapers.vancleefarpels import handle_vancleefarpels
from scrapers.davidyurman import handle_davidyurman
from scrapers.chopard import handle_chopard
from scrapers.jonehardy import handle_jonehardy
from scrapers.anitako import handle_anitako
from scrapers.jennifermeyer import handle_jennifermeyer 
from scrapers.jacquieaiche import handle_jacquieaiche
from scrapers.jacobandco import handle_jacobandco
from scrapers.ferkos import handle_ferkos
from scrapers.heartsonfire import handle_heartsonfire

################################################## 26 /04 ###########################################################
from scrapers.chanel import handle_chanel
from scrapers.buccellati import handle_buccellati
from scrapers.harrywinston import handle_harrywinston

from scrapers.jadetrau import handle_jadetrau
from scrapers.vrai import handle_vrai
from scrapers.stephaniegottlieb import handle_stephaniegottlieb
from scrapers.marcobicego import handle_marcobicego
from scrapers.ringconcierge import handle_ringconcierge
from scrapers.eastwestgemco import handle_eastwestgemco
from scrapers.facets import handle_facets
from scrapers.birks import handle_birks
from scrapers.boochier import handle_boochier

############################################# 28/04  ########################################################
from scrapers.graff import handle_graff
from scrapers.mejuri import handle_mejuri
from scrapers.boucheron import handle_boucheron
from scrapers.chaumet import handle_chaumet
from scrapers.brilliantearth import handle_brilliantearth
from scrapers.forevermark import handle_forevermark
from scrapers.louisvuitton import handle_louisvuitton

from scrapers.piaget import handle_piaget
from scrapers.harrods import handle_harrods
from scrapers.cartier import handle_cartier
from scrapers.bulgari import handle_bulgari
from scrapers.laurenbjewelry1 import handle_laurenbjewelry1
from scrapers.ajaffe import handle_ajaffe


#############################################################################################################

import asyncio
from flask_cors import CORS
from utils import get_public_ip,log_event
from limit_checker import check_monthly_limit
import json
from database import reset_scraping_limit,get_scraping_settings,get_all_scraped_products

from dotenv import load_dotenv
load_dotenv
app = Flask(__name__)
CORS(app)
#############################################################################################################
import logging
import os

DB_CONFIG = {
    "server": os.getenv("DB_SERVER"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME"),
}

os.makedirs("logs", exist_ok=True)

# File to store request count
request_count_file = "logs/proxy_request_count.txt"

# Read request count from file or initialize it
if os.path.exists(request_count_file):
    with open(request_count_file, "r") as f:
        try:
            request_count = int(f.read().strip())
        except ValueError:
            request_count = 0
else:
    request_count = 0

def log_and_increment_request_count():
    """Increment and log the number of requests made via proxy."""
    global request_count
    request_count += 1
    with open(request_count_file, "w") as f:
        f.write(str(request_count))
    logging.info(f"Total requests via proxy: {request_count}")


        
#############################################################################################################
# Load JSON data
def load_websites():
    with open("retailer.json", "r") as file:
        return json.load(file)["websites"]

@app.route("/")
def main():
    websites = load_websites()
    
    return render_template("main.html", websites=websites)

@app.route('/fetch', methods=['POST'])
def fetch_data():
    
    # is_valid, message = check_proxies()
    # if not is_valid:
    #     print("OUT VALIDATION CODE")
    #     print(message)
    #     return jsonify({"errormsg": "Proxy validation failed"}), 502
    
    # Check the daily limit
    if not check_monthly_limit():
        return jsonify({"errormsg": "Daily limit reached. Scraping is disabled."}), 400
   
    # Get URL and pagination details
    url = request.form.get('url')
    max_pages = int(request.form.get('maxPages', 1))  # Ensure max_pages is an integer

    # print("Final URL:", final_url)
    domain = urlparse(url).netloc.lower()
    
    # scrape_id = generate_unique_id(url)
    # insert_scrape_log(scrape_id, url, 'active')
    print(domain)
    logging.info(f"Processing request for domain: {domain}")

    # Increment and log request count
    log_and_increment_request_count()

    # Check domain and call corresponding handler function
    if 'www.jared.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jared(url, max_pages))    
    elif 'www.kay.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_kay(url, max_pages))    
    elif 'www.fhinds.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_fhinds(url, max_pages))
    elif 'www.ernestjones.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_ernest_jones(url, max_pages))
    elif 'www.gabrielny.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_gabriel(url, max_pages)) 
    elif 'www.hsamuel.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_h_samuel(url, max_pages)) 
    elif 'www.tiffany.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_tiffany(url, max_pages)) 
    elif 'www.shaneco.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_shane_co(url, max_pages))
#======================================================================#
    elif 'www.kayoutlet.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_kayoutlet(url, max_pages)) 
    elif 'www.zales.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_zales(url, max_pages))       
    elif 'www.helzberg.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_helzberg(url, max_pages))
    elif 'www.ross-simons.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_rosssimons(url, max_pages))
    elif 'www.peoplesjewellers.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_peoplesjewellers(url, max_pages))  
    elif 'www.fraserhart.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_fraserhart(url, max_pages)) 
    elif 'www.fields.ie' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_fields(url, max_pages))
    elif 'www.warrenjames.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_warrenjames(url, max_pages))
    elif 'www.goldsmiths.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_goldsmiths(url, max_pages))
    elif 'www.thediamondstore.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_thediamondstore(url, max_pages))
    elif 'www.prouds.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_prouds(url, max_pages)) 
    elif 'goldmark.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_goldmark(url, max_pages))
    elif 'www.anguscoote.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_anguscoote(url, max_pages))   
    elif 'bash.com' in domain:  
        base64_encoded, filename, file_path = asyncio.run(handle_bash(url, max_pages)) 
    elif 'www.shiels.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_shiels(url, max_pages)) 
    elif 'mazzucchellis.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mazzucchellis(url, max_pages)) 
    elif 'hoskings.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_hoskings(url, max_pages)) 
    elif 'www.hardybrothers.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_hardybrothers(url, max_pages))
    elif 'www.zamels.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_zamels(url, max_pages))
    elif 'www.wallacebishop.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_wallacebishop(url, max_pages)) 
    elif 'www.bevilles.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_bevilles(url, max_pages))    
    elif 'www.michaelhill.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_michaelhill(url, max_pages))
    elif 'www.apart.eu' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_apart(url, max_pages))
    elif 'www.macys.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_macys(url, max_pages))
    elif 'www.jcpenney.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jcpenney(url, max_pages))
    elif 'www.fredmeyerjewelers.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_fredmeyer(url, max_pages))
    elif 'www.beaverbrooks.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_beaverbrooks(url, max_pages)) 
        
######################################### 21/04 ####################################################################                                                                                          
    elif 'www.finks.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_finks(url, max_pages))  
    elif 'smilingrocks.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_smilingrocks(url, max_pages))
    elif 'www.bluenile.com' in domain: 
        base64_encoded, filename, file_path = asyncio.run(handle_bluenile(url, max_pages)) 
    elif 'www.benbridge.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_benbridge(url, max_pages)) 
    elif 'www.hannoush.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_hannoush(url, max_pages)) 
    elif 'www.jcojewellery.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jcojewellery(url, max_pages))
    elif 'www.77diamonds.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_77diamonds(url, max_pages))
    elif 'www.reeds.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_reeds(url, max_pages))
    elif 'www.walmart.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_walmart(url, max_pages))     
############################################# 22/04 ################################################################               
    elif 'armansfinejewellery.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_armansfinejewellery(url, max_pages)) 
    elif 'jacquefinejewellery.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jacquefinejewellery(url, max_pages))
    elif 'medleyjewellery.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_medleyjewellery(url, max_pages))
    elif 'cullenjewellery.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_cullenjewellery(url, max_pages)) 
    elif 'www.grahams.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_grahams(url, max_pages))
    elif 'www.larsenjewellery.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_larsenjewellery(url, max_pages))  
    elif 'ddsdiamonds.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_ddsdiamonds(url, max_pages))
    elif 'www.garenjewellery.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_garenjewellery(url, max_pages))
    elif 'stefandiamonds.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_stefandiamonds(url, max_pages))
    elif 'www.goodstoneinc.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_goodstoneinc(url, max_pages))                             
    elif 'natashaschweitzer.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_natasha(url, max_pages))
    elif 'www.sarahandsebastian.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_sarahandsebastian(url, max_pages))
    elif 'tmcfinejewellers.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_moissanite(url, max_pages))
    elif 'diamondcollective.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_diamondcollection(url, max_pages))
    elif 'cushlawhiting.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_cushlawhiting(url, max_pages))
    elif 'cerrone.com.au' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_cerrone(url, max_pages))     
#############################################################################################################
    elif 'www.briju.pl' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_briju(url, max_pages))
    elif 'www.histoiredor.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_histoiredor(url, max_pages))
    elif 'www.marc-orian.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_marcorian(url, max_pages))
    elif 'www.klenotyaurum.cz' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_klenotyaurum(url, max_pages))       
    elif 'www.stroilioro.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_stroilioro(url, max_pages)) 
    # elif 'bash.com' in domain:
    #     base64_encoded, filename, file_path = asyncio.run(handle_americanswiss(url, max_pages))  
    elif 'mariemas.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mariemass(url, max_pages))
    elif 'mattioli.it' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mattioli(url, max_pages))
    elif 'www.pomellato.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_pomellato(url, max_pages))
    elif 'www.dior.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_dior(url, max_pages))
        
                        
########################################### 24/07 ################################################################## 
    elif 'www.diamondsfactory.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_diamondsfactory(url, max_pages)) 
    elif 'www.davidmarshalllondon.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_davidmarshalllondon(url, max_pages))
    elif 'www.monicavinader.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_monicavinader(url, max_pages))        
    elif 'www.boodles.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_boodles(url, max_pages))
    elif 'www.maria-black.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mariablack(url, max_pages))    
    elif 'www.londonjewelers.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_londonjewelers(url, max_pages))
    elif 'fernandojorge.co.uk' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_fernandojorge(url, max_pages)) 
    elif 'us.pandora.net' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_pandora(url, max_pages)) 
    elif 'www.daisyjewellery.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_daisyjewellery(url, max_pages)) 
    elif 'www.missoma.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_missoma(url, max_pages)) 
    elif 'bybonniejewelry.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_bybonniejewelry(url, max_pages))
    elif 'mateonewyork.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mateo(url, max_pages))
    elif 'edgeofember.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_edgeofember(url, max_pages))
    elif 'www.astleyclarke.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_astleyclarke(url, max_pages))  
################################################ 25/04 ############################################################# 
    elif 'www.tacori.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_tacori(url, max_pages))
    elif 'www.vancleefarpels.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_vancleefarpels(url, max_pages))
    elif 'www.davidyurman.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_davidyurman(url, max_pages))
    elif 'www.chopard.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_chopard(url, max_pages)) 
    elif "johnhardy.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jonehardy(url, max_pages))
    elif "www.anitako.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_anitako(url, max_pages))
    elif "jennifermeyer.com" in domain: 
        base64_encoded, filename, file_path = asyncio.run(handle_jennifermeyer(url, max_pages))
    elif "jacquieaiche.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jacquieaiche(url, max_pages))
    elif "jacobandco.shop" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jacobandco(url, max_pages))
    elif "ferkosfinejewelry.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_ferkos(url, max_pages))
    elif "www.heartsonfire.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_heartsonfire(url, max_pages))
                                             
################################################### 26 /04 ##########################################################
    elif 'www.chanel.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_chanel(url, max_pages)) 
    elif 'www.buccellati.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_buccellati(url, max_pages))
    elif 'www.harrywinston.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_harrywinston(url, max_pages))  
    
    elif "jadetrau.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_jadetrau(url, max_pages))
    elif "www.vrai.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_vrai(url, max_pages))
    elif "stephaniegottlieb.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_stephaniegottlieb(url, max_pages))
    elif "marcobicego.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_marcobicego(url, max_pages))
    elif "ringconcierge.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_ringconcierge(url, max_pages))
    elif "eastwestgemco.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_eastwestgemco(url, max_pages))
    elif "64facets.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_facets(url, max_pages))
    elif "boochier.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_boochier(url, max_pages))
    elif "www.birks.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_birks(url, max_pages))
    
           
############################################# 28/04  ################################################################ 
    elif 'www.graff.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_graff(url, max_pages))
    elif 'mejuri.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_mejuri(url, max_pages))  
    elif 'www.boucheron.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_boucheron(url, max_pages)) 
    elif 'www.chaumet.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_chaumet(url, max_pages)) 
    elif 'www.brilliantearth.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_brilliantearth(url, max_pages))
    elif 'www.forevermark.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_forevermark(url, max_pages))
    elif 'eu.louisvuitton.com' in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_louisvuitton(url, max_pages))
    elif "www.piaget.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_piaget(url, max_pages))
    elif "www.harrods.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_harrods(url, max_pages))
    elif "www.cartier.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_cartier(url, max_pages))

    elif "www.bulgari.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_bulgari(url, max_pages))
    elif "www.laurenbjewelry.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_laurenbjewelry1(url, max_pages))
    elif "ajaffe.com" in domain:
        base64_encoded, filename, file_path = asyncio.run(handle_ajaffe(url, max_pages))
    
                            
#############################################################################################################
    else:
        log_event(f"Unknown website attempted: {domain}")
        return jsonify({"error": "Unknown website"}), 500
    
    # Return file download link or error message
    if filename:
        # update_scrape_status(scrape_id, 'inactive')
        log_event(f"Successfully scraped {domain}. File generated: {filename}")
        return jsonify({'file': base64_encoded, 'filename': filename, 'filepath': file_path}),200
    else:
        # update_scrape_status(scrape_id, 'error')
        print("output")
        log_event(f"Scraping failed for {domain}")
        return jsonify({"error": "Failed"}), 800



#############################################################################################################

@app.route("/reset-limit", methods=["GET"])
def reset_limit_route():
    result = reset_scraping_limit()
    return (jsonify(result), 200) if not result.get("error") else (jsonify(result), 500)


@app.route("/get_data")
def get_data():
    return jsonify(get_scraping_settings())



@app.route("/get_products", methods=["GET"])
def get_products():
    return jsonify(get_all_scraped_products())



@app.route("/productview")
def productview():
    
    products1 = get_all_scraped_products()
    # print(products)
    # print(type(products))
    return render_template("product_view.html", products1=products1)

#############################################################################################################


@app.route('/report', methods=['POST'])
def generate_report():
    data = request.get_json()
    selected_date = data.get('date')
    selected_header = data.get('header')

    print(f"Received date: {selected_date}, header: {selected_header}")

    try:
        # Connect to SQL Server
        conn = pymssql.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Query to fetch data - verify this matches your table structure
        query = """
        SELECT CurrentDate, Header, ProductName, ImagePath, Kt, Price, TotalDiaWt, Time, AdditionalInfo
        FROM IBM_Algo_Webstudy_Products
        WHERE CONVERT(date, CurrentDate) = %s AND Header = %s
        """
        cursor.execute(query, (selected_date, selected_header))
        records = cursor.fetchall()
        
        if not records:
            return jsonify({"error": "No records found for the selected date and header"}), 404

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Report"

        # Define headers and set column widths
        headers = ["Current Date", "Header", "Product Name", "Image", "Kt", "Price", 
                   "Total Dia wt", "Time", "ImagePath", "Additional Info"]
        ws.append(headers)
        
        column_widths = {'A': 15, 'B': 20, 'C': 25, 'D': 15, 'E': 10, 
                        'F': 15, 'G': 15, 'H': 15, 'I': 40, 'J': 40}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Create temporary directory for image processing
        with tempfile.TemporaryDirectory() as temp_dir:
            # Process records
            for row_idx, record in enumerate(records, start=2):
                try:
                    # Unpack all 9 values at once
                    (current_date, header, product_name, image_path, 
                     kt, price, total_dia_wt, time, additional_info) = record
                    
                    # Add data row
                    ws.append([
                        current_date, header, product_name, 
                        '',  # Placeholder for image
                        kt, price, total_dia_wt, time, 
                        image_path, additional_info
                    ])

                    # Set row height for image rows
                    ws.row_dimensions[row_idx].height = 75

                    # Handle image if path exists
                    if image_path and os.path.exists(image_path):
                        try:
                            # Create unique temp filename
                            temp_img_path = os.path.join(temp_dir, f"img_{row_idx}.jpg")
                            
                            # Process and save image
                            with Image.open(image_path) as img:
                                img = img.convert("RGB")  # <-- This line is essential
                                img.thumbnail((130, 130))
                                img.save(temp_img_path, format="JPEG")  # Save as JPEG explicitly
                            
                            # Add to Excel
                            excel_img = XLImage(temp_img_path)
                            excel_img.anchor = f'D{row_idx}'
                            ws.add_image(excel_img)
                        except Exception as img_error:
                            print(f"Error processing image {image_path}: {img_error}")
                            ws.cell(row=row_idx, column=4, value="Image Error")
                except Exception as row_error:
                    print(f"Error processing row {row_idx}: {row_error}")
                    continue

            # Save workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

        # Close database connection
        cursor.close()
        conn.close()

        # Create response
        response = send_file(
            output,
            as_attachment=True,
            download_name=f"report_{selected_date}_{selected_header}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        return response
    except Exception as e:
        print(f"Error generating report: {e}")
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500



#############################################################################################################


def clean_price(price_str):
    """
    Extracts currency and sale/original prices from strings like:
    '$2,649.99 offer of 50% $5,299.99', '45 EUR', 'USD 199.00', '$149.99', '150.99 USD', etc.
    Returns (sale_price, original_price, currency)
    Ignores percentage discounts like '50% off'.
    """
    if not price_str:
        return 0.0, 0.0, ""

    # Normalize input
    price_str = str(price_str).replace(",", "").upper()

    # Remove percentages (e.g. '50%', '30% OFF', etc.)
    price_str = re.sub(r'\d+%(\s*OFF)?', '', price_str)

    # Currency matching pattern
    currency_pattern = r'(?:([€£$₹])\s?|(\bUSD|\bEUR|\bGBP|\bCAD|\bAUD|\bINR)\s?)?(\d+(?:\.\d+)?)(?:\s?(USD|EUR|GBP|CAD|AUD|INR))?'

    matches = re.findall(currency_pattern, price_str)

    prices = []
    currencies = []

    for symbol1, code1, amount, code2 in matches:
        currency = symbol1 or code1 or code2
        if currency:
            currencies.append(currency)
        prices.append(float(amount))

    # Choose most frequent currency if available
    currency = Counter(currencies).most_common(1)[0][0] if currencies else ""

    if len(prices) >= 2:
        return prices[0], prices[1], currency
    elif len(prices) == 1:
        return prices[0], 0.0, currency
    else:
        return 0.0, 0.0, currency

# Determine how to format based on symbol/code position
def format_currency(value, currency):
    if not value:
        return ""
    if currency in ['£', '$', '€', '₹']:
        return f"{currency}{value:.2f}"
    elif currency:
        return f"{value:.2f} {currency}"
    else:
        return f"{value:.2f}"


def safe_str(value):
    value = str(value).strip() if value else ""
    return "'" + value if value.startswith("=") else value


@app.route('/reportsummery', methods=['POST'])
def reportsummary():
    data = request.get_json()
    selected_date = data.get('date')
    selected_header = data.get('header')

    print(f"Received date: {selected_date}, header: {selected_header}")

    temp_dir = tempfile.mkdtemp()
    conn = None
    cursor = None

    try:
        conn = pymssql.connect(**DB_CONFIG)
        cursor = conn.cursor()

        query = """
        SELECT ProductName, ImagePath, Kt, Price, TotalDiaWt
        FROM IBM_Algo_Webstudy_Products
        WHERE CONVERT(date, CurrentDate) = %s AND Header = %s
        """
        cursor.execute(query, (selected_date, selected_header))
        records = cursor.fetchall()

        if not records:
            print("No records found.")
            return jsonify({"message": "No records found for the given date and header."}), 404

        # Excel setup
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visual Product Summery Report"

        # Set column widths
        for i in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25

        # Define styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        products_per_row = 2
        product_height = 8  # 1 image + 6 details + 1 spacing

        for idx, (product_name, image_path, kt, price, total_dia_wt) in enumerate(records):
            try:
                row_offset = (idx // products_per_row) * product_height + 1
                position_in_row = idx % products_per_row
                col = (position_in_row * 2) + 1
                col_val = col + 1
                col_letter = openpyxl.utils.get_column_letter(col)

                product_name = safe_str(product_name)
                kt = safe_str(kt)
                total_dia_wt = safe_str(total_dia_wt)
                sale_price_value, price_value, currency = clean_price(price)
                sale_price_str = format_currency(sale_price_value, currency)
                price_str = format_currency(price_value, currency)

                # Add Image
                if image_path and os.path.exists(image_path):
                    try:
                        temp_img_path = os.path.join(temp_dir, f"img_{idx}.jpg")
                        with Image.open(image_path) as img:
                            img = img.convert("RGB")  # <-- This line is essential
                            img.thumbnail((130, 130))
                            img.save(temp_img_path, format="JPEG")  # Save as JPEG explicitly
                            
                        img_for_excel = XLImage(temp_img_path)
                        img_for_excel.anchor = f'{col_letter}{row_offset}'
                        ws.add_image(img_for_excel)
                        ws.row_dimensions[row_offset].height = 110
                        ws.merge_cells(start_row=row_offset, start_column=col,
                                       end_row=row_offset, end_column=col_val)
                    except Exception as img_err:
                        print(f"Image error for product '{product_name}': {img_err}")
                        ws.cell(row=row_offset, column=col, value="Image error")
                        ws.merge_cells(start_row=row_offset, start_column=col,
                                       end_row=row_offset, end_column=col_val)
                else:
                    ws.cell(row=row_offset, column=col, value="No image")
                    ws.merge_cells(start_row=row_offset, start_column=col,
                                   end_row=row_offset, end_column=col_val)

                # Product Details
                labels = ["Product Name", "Kt", "Sale Price", "Price", "Total Dia wt", "Gold wt"]
                values = [
                    product_name,
                    kt,
                    sale_price_str,
                    price_str,
                    total_dia_wt,
                    ""
                ]
                for i, (label, val) in enumerate(zip(labels, values), start=1):
                    label_cell = ws.cell(row=row_offset + i, column=col, value=label)
                    val_cell = ws.cell(row=row_offset + i, column=col_val, value=val)

                    label_cell.font = bold_font
                    label_cell.alignment = left_align
                    label_cell.border = thin_border

                    val_cell.alignment = left_align
                    val_cell.border = thin_border

            except Exception as row_err:
                print(f"Error processing record {idx + 1}: {row_err}")
                ws.cell(row=row_offset + 1, column=col, value="ERROR")

        # Save workbook to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"report_{selected_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error generating report: {e}")
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500

   
#############################################################################################################
@app.route("/summery")
def summery():
    
    return render_template("summery.html")

@app.route("/api/category-summary")
def category_summary():
    try:
        with pymssql.connect(**DB_CONFIG) as conn:
            with conn.cursor(as_dict=True) as cursor:
                # Step 1: Get distinct portals from the Header field
                cursor.execute("""
                    SELECT DISTINCT
                        CASE
                            WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                            WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                            WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                            WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                            WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                            ELSE ''
                        END AS Portal
                    FROM IBM_Algo_Webstudy_Products
                """)
                
                portals = [row['Portal'] for row in cursor.fetchall() if row['Portal']]

                if not portals:
                    return jsonify({"success": True, "data": []})

                # Step 2: Construct dynamic SQL components
                columns = ', '.join(f"[{portal}]" for portal in portals)
                columns_with_isnull = ', '.join(f"ISNULL([{portal}], 0) AS [{portal}]" for portal in portals)
                columns_for_total = ' + '.join(f"ISNULL([{portal}], 0)" for portal in portals)

                # Step 3: Construct the final dynamic SQL
                dynamic_query = f"""
                    WITH Extracted AS (
                        SELECT
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                                ELSE ''
                            END AS Portal,
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('–', Header) - 1)))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('|', Header) - 1)))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' I ', Header) - 1)))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' / ', Header) - 1)))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' - ', Header) - 1)))
                                ELSE ''
                            END AS Category
                        FROM IBM_Algo_Webstudy_Products
                    ),
                    Counts AS (
                        SELECT Category, Portal, COUNT(*) AS Total
                        FROM Extracted
                        WHERE Portal <> ''
                        GROUP BY Category, Portal
                    ),
                    Pivoted AS (
                        SELECT Category, {columns}
                        FROM Counts
                        PIVOT (
                            SUM(Total) FOR Portal IN ({columns})
                        ) AS p
                    )
                    SELECT 
                        Category,
                        {columns_for_total} AS TOTAL,
                        {columns_with_isnull}
                    FROM Pivoted
                    ORDER BY Category;
                """

                # Step 4: Execute the final query
                cursor.execute(dynamic_query)
                rows = cursor.fetchall()

                return jsonify({"success": True, "data": rows})

    except pymssql.Error as e:
        return jsonify({"success": False, "error": str(e)})


    
#-----------------------------------------------------------------------------------------------------#    
@app.route("/api/diawt-summary")
def diawt_summary():
    try:
        with pymssql.connect(**DB_CONFIG) as conn:
            with conn.cursor(as_dict=True) as cursor:
                # Step 1: Get distinct portals dynamically
                cursor.execute("""
                    SELECT DISTINCT
                        CASE
                            WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                            WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                            WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                            WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                            WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                            ELSE NULL
                        END AS portal
                    FROM IBM_Algo_Webstudy_Products
                    WHERE Header IS NOT NULL AND Header <> ''
                """)
                portals = sorted({row['portal'] for row in cursor if row['portal']})

                if not portals:
                    return jsonify({"success": False, "error": "No portal headers found."})

                # Prepare dynamic columns for PIVOT
                cols = ','.join(f"[{p}]" for p in portals)
                cols_isnull_sum = ' + '.join(f"ISNULL([{p}], 0)" for p in portals)
                cols_isnull_select = ', '.join(f"ISNULL([{p}], 0) AS [{p}]" for p in portals)

                # Step 2: Dynamic SQL query without Unknown row
                dynamic_sql = f"""
                    WITH Cleaned AS (
                        SELECT
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                                ELSE NULL
                            END AS portal,
                            LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(TotalDiaWt, 'ct', ''), 'tw', ''), ' ', ''), '–', '-'), '--', '-')) AS DiaWtStr
                        FROM IBM_Algo_Webstudy_Products
                        WHERE Header IS NOT NULL AND TotalDiaWt IS NOT NULL AND TotalDiaWt <> ''
                    ),
                    Converted AS (
                        SELECT
                            portal,
                            CASE
                                WHEN DiaWtStr LIKE '%-%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                    + (
                                        TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('-', DiaWtStr) + 1, CHARINDEX('/', DiaWtStr) - CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                        /
                                        TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                    )
                                WHEN DiaWtStr LIKE '%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('/', DiaWtStr) - 1) AS FLOAT)
                                    /
                                    TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                WHEN DiaWtStr LIKE '.%' THEN TRY_CAST('0' + DiaWtStr AS FLOAT)
                                ELSE TRY_CAST(DiaWtStr AS FLOAT)
                            END AS DiaWtClean
                        FROM Cleaned
                        WHERE portal IS NOT NULL
                    ),
                    Binned AS (
                        SELECT
                            portal,
                            CASE
                                WHEN DiaWtClean < 0.10 THEN 'Below 0.10ct'
                                WHEN DiaWtClean >= 0.10 AND DiaWtClean < 0.25 THEN '0.10ct - 0.24ct'
                                WHEN DiaWtClean >= 0.25 AND DiaWtClean < 0.50 THEN '0.25ct - 0.49ct'
                                WHEN DiaWtClean >= 0.50 AND DiaWtClean < 0.75 THEN '0.50ct - 0.74ct'
                                WHEN DiaWtClean >= 0.75 AND DiaWtClean < 1.00 THEN '0.75ct - 0.99ct'
                                WHEN DiaWtClean >= 1.00 AND DiaWtClean < 1.50 THEN '1.00ct - 1.49ct'
                                WHEN DiaWtClean >= 1.50 AND DiaWtClean <= 2.00 THEN '1.50ct - 2.00ct'
                                WHEN DiaWtClean > 2.00 THEN 'Above 2.00ct'
                                ELSE NULL  -- Exclude Unknown by returning NULL
                            END AS WeightRange
                        FROM Converted
                        WHERE DiaWtClean IS NOT NULL  -- Exclude unconvertible values
                    ),
                    Counts AS (
                        SELECT WeightRange, portal, COUNT(*) AS cnt
                        FROM Binned
                        WHERE WeightRange IS NOT NULL  -- Exclude NULL weight ranges
                        GROUP BY WeightRange, portal
                    ),
                    Pivoted AS (
                        SELECT
                            WeightRange,
                            {cols},
                            {cols_isnull_sum} AS TOTAL
                        FROM Counts
                        PIVOT (
                            SUM(cnt) FOR portal IN ({cols})
                        ) AS P
                    ),
                    -- Calculate grand totals per portal
                    PortalTotals AS (
                        SELECT 
                            portal,
                            SUM(cnt) AS portal_total
                        FROM Counts
                        GROUP BY portal
                    ),
                    PivotedTotals AS (
                        SELECT
                            'Total' AS WeightRange,
                            {cols},
                            {cols_isnull_sum} AS TOTAL
                        FROM PortalTotals
                        PIVOT (
                            SUM(portal_total) FOR portal IN ({cols})
                        ) AS P
                    ),
                    -- Combine category rows with total row
                    Combined AS (
                        SELECT * FROM Pivoted
                        UNION ALL
                        SELECT * FROM PivotedTotals
                    )
                    SELECT
                        WeightRange AS [DIA WTS],
                        {cols_isnull_select},
                        ISNULL(TOTAL, 0) AS TOTAL
                    FROM Combined
                    ORDER BY
                        CASE WeightRange
                            WHEN 'Below 0.10ct' THEN 1
                            WHEN '0.10ct - 0.24ct' THEN 2
                            WHEN '0.25ct - 0.49ct' THEN 3
                            WHEN '0.50ct - 0.74ct' THEN 4
                            WHEN '0.75ct - 0.99ct' THEN 5
                            WHEN '1.00ct - 1.49ct' THEN 6
                            WHEN '1.50ct - 2.00ct' THEN 7
                            WHEN 'Above 2.00ct' THEN 8
                            ELSE 9  -- Total row
                        END
                """

                cursor.execute(dynamic_sql)
                rows = cursor.fetchall()
                return jsonify({"success": True, "data": rows})

    except pymssql.Error as e:
        return jsonify({"success": False, "error": str(e)})
#-----------------------------------------------------------------------------------------------------#
@app.route("/api/combined-summary")
def combined_summary():
    try:
        with pymssql.connect(**DB_CONFIG) as conn:
            with conn.cursor(as_dict=True) as cursor:
                cursor.execute("""
                    WITH Cleaned AS (
                        SELECT
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                                ELSE NULL
                            END AS portal,
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('–', Header) - 1)))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('|', Header) - 1)))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' I ', Header) - 1)))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' / ', Header) - 1)))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' - ', Header) - 1)))
                                ELSE LTRIM(RTRIM(Header))
                            END AS Category,
                            LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(TotalDiaWt, 'ct', ''), 'tw', ''), ' ', ''), '–', '-'), '--', '-')) AS DiaWtStr
                        FROM IBM_Algo_Webstudy_Products
                        WHERE Header IS NOT NULL AND Header <> ''
                            AND TotalDiaWt IS NOT NULL AND TotalDiaWt <> ''
                    ),
                    Converted AS (
                        SELECT
                            portal,
                            Category,
                            CASE
                                WHEN DiaWtStr LIKE '%-%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                    + (
                                        TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('-', DiaWtStr) + 1, CHARINDEX('/', DiaWtStr) - CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                        /
                                        TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                    )
                                WHEN DiaWtStr LIKE '%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('/', DiaWtStr) - 1) AS FLOAT)
                                    / TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                WHEN DiaWtStr LIKE '.%' THEN TRY_CAST('0' + DiaWtStr AS FLOAT)
                                ELSE TRY_CAST(DiaWtStr AS FLOAT)
                            END AS DiaWtClean
                        FROM Cleaned
                        WHERE portal IS NOT NULL AND Category IS NOT NULL
                    )
                    SELECT
                        Category,
                        SUM(CASE WHEN DiaWtClean < 0.10 THEN 1 ELSE 0 END) AS [Below 0.10ct],
                        SUM(CASE WHEN DiaWtClean >= 0.10 AND DiaWtClean < 0.25 THEN 1 ELSE 0 END) AS [0.10ct - 0.24ct],
                        SUM(CASE WHEN DiaWtClean >= 0.25 AND DiaWtClean < 0.50 THEN 1 ELSE 0 END) AS [0.25ct - 0.49ct],
                        SUM(CASE WHEN DiaWtClean >= 0.50 AND DiaWtClean < 0.75 THEN 1 ELSE 0 END) AS [0.50ct - 0.74ct],
                        SUM(CASE WHEN DiaWtClean >= 0.75 AND DiaWtClean < 1.00 THEN 1 ELSE 0 END) AS [0.75ct - 0.99ct],
                        SUM(CASE WHEN DiaWtClean >= 1.00 AND DiaWtClean < 1.50 THEN 1 ELSE 0 END) AS [1.00ct - 1.49ct],
                        SUM(CASE WHEN DiaWtClean >= 1.50 AND DiaWtClean <= 2.00 THEN 1 ELSE 0 END) AS [1.50ct - 2.00ct],
                        SUM(CASE WHEN DiaWtClean > 2.00 THEN 1 ELSE 0 END) AS [Above 2.00ct],
                        SUM(CASE WHEN DiaWtClean IS NOT NULL THEN 1 ELSE 0 END) AS Total
                    FROM Converted
                    GROUP BY portal, Category
                    ORDER BY portal, Category;
                """)
                rows = cursor.fetchall()
                return jsonify({"success": True, "data": rows})

    except pymssql.Error as e:
        return jsonify({"success": False, "error": str(e)})

#-----------------------------------------------------------------------------------------------------#
@app.route("/api/compiled-data")
def compiled_data():
    try:
        with pymssql.connect(**DB_CONFIG) as conn:
            with conn.cursor(as_dict=True) as cursor:
                cursor.execute("""
                    WITH Cleaned AS (
                        SELECT
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('–', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX('|', Header) + 1, LEN(Header))))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' I ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' / ', Header) + 3, LEN(Header))))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, CHARINDEX(' - ', Header) + 3, LEN(Header))))
                                ELSE NULL
                            END AS portal,
                            CASE
                                WHEN CHARINDEX('–', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('–', Header) - 1)))
                                WHEN CHARINDEX('|', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX('|', Header) - 1)))
                                WHEN CHARINDEX(' I ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' I ', Header) - 1)))
                                WHEN CHARINDEX(' / ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' / ', Header) - 1)))
                                WHEN CHARINDEX(' - ', Header) > 0 THEN LTRIM(RTRIM(SUBSTRING(Header, 1, CHARINDEX(' - ', Header) - 1)))
                                ELSE LTRIM(RTRIM(Header))
                            END AS Category,
                            LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(TotalDiaWt, 'ct', ''), 'tw', ''), ' ', ''), '–', '-'), '--', '-')) AS DiaWtStr
                        FROM IBM_Algo_Webstudy_Products
                        WHERE Header IS NOT NULL AND Header <> ''
                            AND TotalDiaWt IS NOT NULL AND TotalDiaWt <> ''
                    ),
                    CleanedPortal AS (
                        SELECT
                            -- Extract only brand name from portal
                            CASE 
                                -- Remove page references
                                WHEN portal LIKE '% - Page %' 
                                    THEN LEFT(portal, CHARINDEX(' - Page ', portal) - 1)
                                WHEN portal LIKE '% – Page %' 
                                    THEN LEFT(portal, CHARINDEX(' – Page ', portal) - 1)
                                -- Take last segment after final delimiter
                                WHEN CHARINDEX('|', portal) > 0 
                                    THEN LTRIM(RTRIM(REVERSE(LEFT(REVERSE(portal), CHARINDEX('|', REVERSE(portal)) - 1))))
                                WHEN CHARINDEX('–', portal) > 0 
                                    THEN LTRIM(RTRIM(REVERSE(LEFT(REVERSE(portal), CHARINDEX('–', REVERSE(portal)) - 1))))
                                WHEN CHARINDEX(' - ', portal) > 0 
                                    THEN LTRIM(RTRIM(REVERSE(LEFT(REVERSE(portal), CHARINDEX(' - ', REVERSE(portal)) - 1))))
                                ELSE portal 
                            END AS Customer,
                            Category,
                            DiaWtStr
                        FROM Cleaned
                        WHERE portal IS NOT NULL
                    ),
                    Converted AS (
                        SELECT
                            Customer,
                            Category,
                            CASE
                                WHEN DiaWtStr LIKE '%-%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                    + (
                                        TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('-', DiaWtStr) + 1, CHARINDEX('/', DiaWtStr) - CHARINDEX('-', DiaWtStr) - 1) AS FLOAT)
                                        / TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                    )
                                WHEN DiaWtStr LIKE '%/%' THEN
                                    TRY_CAST(LEFT(DiaWtStr, CHARINDEX('/', DiaWtStr) - 1) AS FLOAT)
                                    / TRY_CAST(SUBSTRING(DiaWtStr, CHARINDEX('/', DiaWtStr) + 1, LEN(DiaWtStr)) AS FLOAT)
                                WHEN DiaWtStr LIKE '.%' THEN TRY_CAST('0' + DiaWtStr AS FLOAT)
                                ELSE TRY_CAST(DiaWtStr AS FLOAT)
                            END AS DiaWtClean
                        FROM CleanedPortal
                    )
                    SELECT
                        Customer,
                        Category,
                        SUM(CASE WHEN DiaWtClean < 0.10 THEN 1 ELSE 0 END) AS [Below 0.10ct],
                        SUM(CASE WHEN DiaWtClean >= 0.10 AND DiaWtClean < 0.25 THEN 1 ELSE 0 END) AS [0.10ct - 0.24ct],
                        SUM(CASE WHEN DiaWtClean >= 0.25 AND DiaWtClean < 0.50 THEN 1 ELSE 0 END) AS [0.25ct - 0.49ct],
                        SUM(CASE WHEN DiaWtClean >= 0.50 AND DiaWtClean < 0.75 THEN 1 ELSE 0 END) AS [0.50ct - 0.74ct],
                        SUM(CASE WHEN DiaWtClean >= 0.75 AND DiaWtClean < 1.00 THEN 1 ELSE 0 END) AS [0.75ct - 0.99ct],
                        SUM(CASE WHEN DiaWtClean >= 1.00 AND DiaWtClean < 1.50 THEN 1 ELSE 0 END) AS [1.00ct - 1.49ct],
                        SUM(CASE WHEN DiaWtClean >= 1.50 AND DiaWtClean <= 2.00 THEN 1 ELSE 0 END) AS [1.50ct - 2.00ct],
                        SUM(CASE WHEN DiaWtClean > 2.00 THEN 1 ELSE 0 END) AS [Above 2.00ct]
                    FROM Converted
                    GROUP BY Customer, Category
                    ORDER BY Customer, Category;

                """)
                rows = cursor.fetchall()
                # print(rows)
                return jsonify({"success": True, "data": rows})

    except pymssql.Error as e:
        return jsonify({"success": False, "error": str(e)})      
#############################################################################################################
if __name__ == '__main__':
    # app.run(debug=True)
    app.run(host="0.0.0.0", port=8000)